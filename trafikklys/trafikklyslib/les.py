"""Leser arbeidsboka og setter sammen det møtevisningen trenger."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from .felles import (ARKNAVN, LYSFARGER, OMRADER, PROFILFARGE_STANDARD, kort_dato, nokkel,
                     rens, statusslag, tekster, tolk_dato, tolk_farge, tolk_lys, tolk_sprak,
                     vis_lys)
from .regneark import KLASSERADER, MOTERADER, OMRADERADER


@dataclass
class Resultat:
    data: dict
    merknader: list = field(default_factory=list)


def _ark(wb, nokkelnavn: str):
    """Finner arket uansett hvilken målform boka ble laget på."""
    for navn in ARKNAVN[nokkelnavn]:
        if navn in wb.sheetnames:
            return wb[navn]
    return None


def _kolonne(ws, kolonne: int, fra: int, antall: int) -> list[str]:
    verdier = []
    for r in range(fra, fra + antall):
        v = rens(ws.cell(row=r, column=kolonne).value)
        if v:
            verdier.append(v)
    return verdier


# Kolonnene slik de kan hete. Første navn er det arbeidsboka bruker; resten gjør
# at et uttrekk fra Microsoft Lists kan limes rett inn, uansett kolonnerekkefølge.
FELT = {
    "innmelding": [
        ("mote", ["Møte", "Møter"]),
        ("klasse", ["Klasse"]),
        ("elev", ["Elev"]),
        ("omrade", ["Område"]),
        ("lys", ["Lys"]),
        ("merknad", ["Merknad", "Kommentar"]),
        ("larer", ["Lærer", "Lærar", "Opprettet av", "Oppretta av", "Created By", "Forfatter"]),
    ],
    "tiltak": [
        ("mote", ["Møte", "Møter"]),
        ("klasse", ["Klasse"]),
        ("elev", ["Elev"]),
        ("omrade", ["Område"]),
        ("tiltak", ["Tiltak"]),
        ("ansvarleg", ["Ansvarlig", "Ansvarleg"]),
        ("frist", ["Frist", "Forfallsdato"]),
        ("status", ["Status"]),
    ],
}


def _kolonnekart(ws, slag: str) -> dict:
    """Finner kolonnene ut fra hoderaden. Faller tilbake til rekkefølgen i arbeidsboka."""
    hode = {}
    for c in range(1, min(ws.max_column, 40) + 1):
        n = nokkel(ws.cell(row=1, column=c).value)
        if n and n not in hode:
            hode[n] = c
    kart = {}
    for i, (felt, navn) in enumerate(FELT[slag]):
        treff = next((hode[nokkel(n)] for n in navn if nokkel(n) in hode), None)
        kart[felt] = treff if treff else i + 1
    return kart


def _rader(ws, kart: dict) -> list[tuple[int, dict]]:
    """Alle rader med noe i, med radnummeret slik det står i Excel."""
    bredde = max(kart.values())
    ut = []
    for rad in ws.iter_rows(min_row=2, min_col=1, max_col=bredde):
        verdier = [c.value for c in rad]
        if not any(rens(v) for v in verdier):
            continue
        ut.append((rad[0].row, {felt: verdier[k - 1] for felt, k in kart.items()}))
    return ut


def les(sti) -> Resultat:
    wb = load_workbook(Path(sti), data_only=True)
    merknader: list[str] = []

    opp = _ark(wb, "ark_oppsett")
    if opp is None:
        raise ValueError("Fant ikke Oppsett-arket. Er dette en trafikklys-arbeidsbok?")

    sprak = tolk_sprak(opp["C6"].value or "bokmål")
    T = tekster(sprak)
    skole = rens(opp["C4"].value) or ("Skulen" if sprak == "nn" else "Skolen")
    skolear = rens(opp["C5"].value)
    profilfarge = tolk_farge(opp["C7"].value, PROFILFARGE_STANDARD)
    logofil = rens(opp["C8"].value)

    klasser = _kolonne(opp, 2, 12, KLASSERADER)
    if not klasser:
        merknader.append("Ingen klasser står i Oppsett. Legg dem inn i kolonnen «Klasser».")

    omrader = []
    for i in range(12, 12 + OMRADERADER):
        navn = rens(opp.cell(row=i, column=6).value)
        if navn:
            omrader.append({"navn": navn, "forklaring": rens(opp.cell(row=i, column=7).value)})
    if not omrader:
        omrader = [{"navn": n, "forklaring": f} for n, f in OMRADER[sprak]]
        merknader.append("Ingen områder står i Oppsett. Bruker standardlista.")

    moter = []
    for i in range(12, 12 + MOTERADER):
        navn = rens(opp.cell(row=i, column=3).value)
        if not navn:
            continue
        dato = tolk_dato(opp.cell(row=i, column=4).value)
        moter.append({"navn": navn, "dato": dato.isoformat() if dato else "",
                      "datotekst": kort_dato(dato), "orden": len(moter)})
    moter.sort(key=lambda m: (m["dato"] or "9999", m["orden"]))
    for i, m in enumerate(moter):
        m["orden"] = i

    # ── Elever per klasse ────────────────────────────────────────
    elever: dict[str, list[str]] = {k: [] for k in klasser}
    elevark = _ark(wb, "elevar_ark")
    if elevark is not None:
        for kol in range(1, elevark.max_column + 1):
            klasse = rens(elevark.cell(row=1, column=kol).value)
            if not klasse:
                continue
            navn = _kolonne(elevark, kol, 2, 60)
            elever.setdefault(klasse, [])
            for n in navn:
                if n not in elever[klasse]:
                    elever[klasse].append(n)
            if klasse not in klasser:
                merknader.append(f"«{klasse}» står på {T['elevar_ark']}-arket, men ikke i Oppsett. "
                                 "Klassen er lagt til.")
                klasser.append(klasse)

    kjente_omrader = {nokkel(o["navn"]): o["navn"] for o in omrader}
    kjente_moter = {nokkel(m["navn"]): m["navn"] for m in moter}
    kjente_klasser = {nokkel(k): k for k in klasser}

    def _finn(verdi, kjente, hva, radnr, legg_til=None):
        """Retter små skrivefeil mot lista, eller legger verdien til og sier fra."""
        tekst = rens(verdi)
        if not tekst:
            return ""
        n = nokkel(tekst)
        if n in kjente:
            return kjente[n]
        merknader.append(f"{hva} rad {radnr}: «{tekst}» sto ikke i Oppsett. Den er tatt med som den er.")
        kjente[n] = tekst
        if legg_til is not None:
            legg_til.append(tekst)
        return tekst

    # ── Innmeldinger ─────────────────────────────────────────────
    innmeldingar = []
    inn = _ark(wb, "ark_innmelding")
    if inn is not None:
        for radnr, v in _rader(inn, _kolonnekart(inn, "innmelding")):
            mote = _finn(v["mote"], kjente_moter, T["ark_innmelding"], radnr)
            klasse = _finn(v["klasse"], kjente_klasser, T["ark_innmelding"], radnr, klasser)
            elev = rens(v["elev"])
            omrade = _finn(v["omrade"], kjente_omrader, T["ark_innmelding"], radnr)
            lys = tolk_lys(v["lys"])
            if not elev:
                merknader.append(f"{T['ark_innmelding']} rad {radnr}: ingen elev. Raden er hoppet over.")
                continue
            if not lys:
                merknader.append(f"{T['ark_innmelding']} rad {radnr}: «{rens(v['lys'])}» er ikke grønt, "
                                 "gult eller rødt. Raden er hoppet over.")
                continue
            if klasse and elev not in elever.get(klasse, []):
                elever.setdefault(klasse, []).append(elev)
                merknader.append(f"{T['ark_innmelding']} rad {radnr}: «{elev}» sto ikke på "
                                 f"{T['elevar_ark']}-arket for {klasse}. Eleven er lagt til.")
            if omrade and omrade not in [o["navn"] for o in omrader]:
                omrader.append({"navn": omrade, "forklaring": ""})
            innmeldingar.append({
                "mote": mote, "klasse": klasse, "elev": elev, "omrade": omrade,
                "lys": lys, "lystekst": vis_lys(lys, sprak),
                "merknad": rens(v["merknad"]), "larer": rens(v["larer"]), "rad": radnr,
            })

    # ── Tiltak ───────────────────────────────────────────────────
    slag = statusslag()
    tiltak = []
    til = _ark(wb, "ark_tiltak")
    if til is not None:
        for radnr, v in _rader(til, _kolonnekart(til, "tiltak")):
            if not rens(v["tiltak"]):
                merknader.append(f"{T['ark_tiltak']} rad {radnr}: ingen tiltakstekst. Raden er hoppet over.")
                continue
            frist = tolk_dato(v["frist"])
            status = rens(v["status"])
            tiltak.append({
                "mote": _finn(v["mote"], kjente_moter, T["ark_tiltak"], radnr),
                "klasse": _finn(v["klasse"], kjente_klasser, T["ark_tiltak"], radnr, klasser),
                "elev": rens(v["elev"]),
                "omrade": _finn(v["omrade"], kjente_omrader, T["ark_tiltak"], radnr),
                "tiltak": rens(v["tiltak"]), "ansvarleg": rens(v["ansvarleg"]),
                "frist": frist.isoformat() if frist else "", "fristtekst": kort_dato(frist),
                "status": status, "apen": slag.get(nokkel(status), "apen") == "apen",
                "rad": radnr,
            })

    if not moter and innmeldingar:
        merknader.append("Ingen møter står i Oppsett. Alt havner under ett navnløst møte.")
        moter = [{"navn": "", "dato": "", "datotekst": "", "orden": 0}]

    data = {
        "skole": skole, "skolear": skolear, "sprak": sprak, "tekst": T,
        "profilfarge": profilfarge, "logofil": logofil, "logo": "",
        "lysfarger": LYSFARGER,
        "lystekster": {k: vis_lys(k, sprak) for k in LYSFARGER},
        "klasser": klasser,
        "omrader": omrader,
        "moter": moter,
        "elever": {k: v for k, v in elever.items() if k in klasser},
        "innmeldingar": innmeldingar,
        "tiltak": tiltak,
    }
    return Resultat(data=data, merknader=merknader)
