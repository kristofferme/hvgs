"""Leser arbeidsboka og setter sammen hver uke. Det som ikke stemmer, blir en merknad."""

from __future__ import annotations

import copy
import datetime as dt
import re
from dataclasses import dataclass, field

from openpyxl import load_workbook

from .felles import (ARKNAVN, DAGER, PROFILFARGE_STANDARD, datospenn, farge_for, nokkel,
                     norsk_dato, rens, tekster, tolk_dag, tolk_dato, tolk_farge, tolk_okt,
                     tolk_sprak, tolk_type, uke_til_mandag, vis_dag)


@dataclass
class Resultat:
    data: dict
    merknader: list[str] = field(default_factory=list)


def _rader(ws, antall_kolonner: int, fra_rad: int = 2):
    for nr, rad in enumerate(
        ws.iter_rows(min_row=fra_rad, max_col=antall_kolonner, values_only=True), start=fra_rad
    ):
        if any(v not in (None, "") for v in rad):
            yield nr, list(rad) + [None] * (antall_kolonner - len(rad))


def _kolonneverdier(ws, kolonne: int, fra_rad: int) -> list[str]:
    ut = []
    for (verdi,) in ws.iter_rows(min_row=fra_rad, min_col=kolonne, max_col=kolonne, values_only=True):
        t = rens(verdi)
        if t:
            ut.append(t)
    return ut


def _ark(wb, nokkelnavn: str):
    """Finner arket enten boka er på bokmål eller nynorsk."""
    for navn in ARKNAVN[nokkelnavn]:
        if navn in wb.sheetnames:
            return wb[navn]
    return None


def les(sti) -> Resultat:
    wb = load_workbook(sti, data_only=True)
    merknader: list[str] = []

    for nokkelnavn, vist in (("ark_oppsett", "Oppsett"), ("ark_timeplan", "Timeplan"), ("ark_uke", "Uke")):
        if _ark(wb, nokkelnavn) is None:
            raise SystemExit(f"Arket «{vist}» mangler i {sti}. Lag en ny arbeidsbok med: ukeplan.py ny")

    opp = _ark(wb, "ark_oppsett")
    sprak = tolk_sprak(opp["C8"].value)
    T = tekster(sprak)
    skole = rens(opp["C4"].value) or T["skole"]
    overskrift = rens(opp["C7"].value) or T["ukeplan"]
    profilfarge = tolk_farge(opp["C9"].value, PROFILFARGE_STANDARD)
    logo = rens(opp["C10"].value)

    forste = tolk_dato(opp["C6"].value)
    uke_celle = rens(opp["C5"].value)
    if forste is None and uke_celle.isdigit():
        forste = uke_til_mandag(int(uke_celle), dt.date.today())
        merknader.append("Mandagsdatoen manglet. Datoene er regnet ut fra ukenummeret.")
    if forste is None:
        i_dag = dt.date.today()
        forste = i_dag - dt.timedelta(days=i_dag.weekday())
        merknader.append("Verken dato eller ukenummer var satt. Bruker inneværende uke.")
    forste -= dt.timedelta(days=forste.weekday())
    standarduke = int(uke_celle) if uke_celle.isdigit() else forste.isocalendar()[1]

    klasser = [k for k in _kolonneverdier(opp, 2, 14) if nokkel(k) != "alle"]
    fagnavn = _kolonneverdier(opp, 5, 13)
    egne_farger = {}
    for i, navn in enumerate(fagnavn):
        hex_ = rens(opp.cell(row=13 + i, column=6).value)
        if hex_.startswith("#") and len(hex_) == 7:
            egne_farger[navn] = hex_

    kjente_klasser = {nokkel(k): k for k in klasser}
    kjente_fag = {nokkel(f): f for f in fagnavn}

    def fest_klasse(verdi, hvor):
        t = rens(verdi)
        if not t or nokkel(t) == "alle":
            return "Alle"
        if nokkel(t) in kjente_klasser:
            return kjente_klasser[nokkel(t)]
        kjente_klasser[nokkel(t)] = t
        klasser.append(t)
        merknader.append(f"{hvor}: klassen «{t}» sto ikke i Oppsett. Den er lagt til.")
        return t

    def fest_fag(verdi, hvor):
        t = rens(verdi)
        if not t:
            return ""
        if nokkel(t) in kjente_fag:
            return kjente_fag[nokkel(t)]
        kjente_fag[nokkel(t)] = t
        fagnavn.append(t)
        merknader.append(f"{hvor}: faget «{t}» sto ikke i Oppsett. Det er lagt til.")
        return t

    laerere = _les_laerere(wb, fest_klasse, fest_fag)
    grunntimeplan = _les_timeplan(_ark(wb, "ark_timeplan"), fest_klasse, fest_fag, laerere, merknader)
    ukeinnhold, beskjeder = _les_innhold(wb, standarduke, fest_klasse, fest_fag, merknader)

    ukenumre = sorted({standarduke} | set(ukeinnhold) | set(beskjeder),
                      key=lambda u: uke_til_mandag(u, forste))
    uker = []
    for nummer in ukenumre:
        mandag = uke_til_mandag(nummer, forste)
        timer = copy.deepcopy(grunntimeplan)
        oppgaver = []
        for nr, rad in ukeinnhold.get(nummer, []):
            _fest_innhold(timer, oppgaver, rad, nr, merknader)
        uker.append({
            "uke": nummer,
            "mandag": mandag.isoformat(),
            "datospenn": datospenn(mandag, mandag + dt.timedelta(days=4)),
            "dager": [
                {
                    "navn": d,
                    "dato": (mandag + dt.timedelta(days=i)).isoformat(),
                    "visning": norsk_dato(mandag + dt.timedelta(days=i)),
                }
                for i, d in enumerate(DAGER)
            ],
            "timer": timer,
            "oppgaver": oppgaver,
            "beskjeder": beskjeder.get(nummer, []),
        })

    brukte: dict[str, str] = {}
    fag_ut = [{"navn": f, "farge": egne_farger.get(f) or farge_for(f, brukte)} for f in fagnavn]

    if not grunntimeplan:
        merknader.append("Timeplanen er tom. Nettsiden viser bare beskjeder og lekser.")
    if not klasser:
        merknader.append("Ingen klasser funnet. Legg dem inn i Oppsett.")

    _vis_pa_malform(uker, sprak)
    data = {
        "skole": skole,
        "overskrift": overskrift,
        "standarduke": standarduke,
        "sprak": sprak,
        "tekst": T,
        "profilfarge": profilfarge,
        "logofil": logo,
        "klasser": klasser,
        "fag": fag_ut,
        "uker": uker,
    }
    return Resultat(data=data, merknader=merknader)


def _vis_pa_malform(uker, sprak: str) -> None:
    """Dagnavnene skrives i den målforma boka bruker."""
    for uke in uker:
        for dag in uke["dager"]:
            dag["navn"] = vis_dag(dag["navn"], sprak)
        for time in uke["timer"]:
            time["dag"] = vis_dag(time["dag"], sprak)
        for oppgave in uke["oppgaver"]:
            oppgave["dag"] = vis_dag(oppgave["dag"], sprak)
            oppgave["frist"] = vis_dag(oppgave["frist"], sprak)


def _les_laerere(wb, fest_klasse, fest_fag) -> dict:
    """(klasse, fag) → lærer. Tom klasse gjelder alle."""
    ut = {}
    ark = _ark(wb, "ark_larere")
    if ark is None:
        return ut
    for nr, (klasse, fag, laerer) in _rader(ark, 3):
        f = fest_fag(fag, f"Lærere rad {nr}")
        if not f:
            continue
        k = fest_klasse(klasse, f"Lærere rad {nr}")
        ut[(nokkel(k), nokkel(f))] = rens(laerer)
    return ut


def _les_timeplan(ws, fest_klasse, fest_fag, laerere, merknader) -> list[dict]:
    """Rutenettet: klassenavn og dagnavn på hoderaden, klokkeslett i kolonne A."""
    timer: list[dict] = []
    klasse = ""
    for nr, rad in _rader(ws, 6, fra_rad=1):
        forste_celle = rens(rad[0])
        if tolk_dag(rad[1]) == "Mandag" and tolk_dag(rad[2]) == "Tirsdag":
            klasse = fest_klasse(forste_celle, f"Timeplan rad {nr}")
            continue
        if not forste_celle:
            if any(rens(x) for x in rad[1:6]):
                merknader.append(
                    f"Timeplan rad {nr}: mangler klokkeslett i kolonne A. Timene der er hoppet over."
                )
            continue
        start, slutt = tolk_okt(forste_celle)
        if not start:
            if any(rens(x) for x in rad[1:6]):
                merknader.append(
                    f"Timeplan rad {nr}: mangler klokkeslett i kolonne A. Timene der er hoppet over."
                )
            continue
        if not klasse:
            merknader.append(f"Timeplan rad {nr}: står over den første klassen. Raden er hoppet over.")
            continue
        for i, dag in enumerate(DAGER):
            fag = fest_fag(rad[1 + i], f"Timeplan rad {nr}")
            if not fag:
                continue
            laerer = laerere.get((nokkel(klasse), nokkel(fag))) or laerere.get(("alle", nokkel(fag))) or ""
            timer.append({
                "klasse": klasse, "dag": dag, "start": start, "slutt": slutt, "fag": fag,
                "laerer": laerer, "tema": "", "lekse": "", "type": "",
            })
    return timer


def _les_innhold(wb, standarduke, fest_klasse, fest_fag, merknader):
    """Uke- og beskjedradene, sortert i hver sin uke."""
    ukeinnhold: dict[int, list] = {}
    for nr, rad in _rader(_ark(wb, "ark_uke"), 8):
        uke, klasse, dag, fag, tema, lekse, frist, type_ = rad
        nummer = _tolk_uke(uke, standarduke, f"Uke rad {nr}", merknader)
        d = tolk_dag(dag)
        if not d and not tolk_dag(frist):
            sto = rens(dag)
            merknader.append(
                f"Uke rad {nr}: «{sto}» er ikke en dag mandag–fredag. Raden er hoppet over."
                if sto else f"Uke rad {nr}: mangler dag. Raden er hoppet over."
            )
            continue
        ukeinnhold.setdefault(nummer, []).append((nr, {
            "klasse": fest_klasse(klasse, f"Uke rad {nr}"),
            "dag": d,
            "fag": fest_fag(fag, f"Uke rad {nr}"),
            "tema": rens(tema),
            "lekse": rens(lekse),
            "frist": tolk_dag(frist),
            "type": tolk_type(type_),
        }))

    beskjeder: dict[int, list] = {}
    beskjedark = _ark(wb, "ark_beskjeder")
    if beskjedark is not None:
        for nr, (uke, klasse, tittel, tekst) in _rader(beskjedark, 4):
            if not rens(tekst) and not rens(tittel):
                continue
            nummer = _tolk_uke(uke, standarduke, f"Beskjeder rad {nr}", merknader)
            beskjeder.setdefault(nummer, []).append({
                "klasse": fest_klasse(klasse, f"Beskjeder rad {nr}"),
                "tittel": rens(tittel),
                "tekst": rens(tekst),
            })
    return ukeinnhold, beskjeder


def _tolk_uke(verdi, standarduke: int, hvor: str, merknader) -> int:
    """Godtar «36», 36 og merkelappen fra nedtrekket: «36 · 31. aug – 4. sep 2026»."""
    tekst = rens(verdi)
    if not tekst:
        return standarduke
    treff = re.match(r"^(\d{1,2})\b", tekst)
    if not treff:
        merknader.append(f"{hvor}: «{tekst}» er ikke et ukenummer. Raden havner i uke {standarduke}.")
        return standarduke
    nummer = int(treff.group(1))
    if not 1 <= nummer <= 53:
        merknader.append(f"{hvor}: uke {nummer} finnes ikke. Raden havner i uke {standarduke}.")
        return standarduke
    return nummer


def _fest_innhold(timer, oppgaver, rad, nr, merknader) -> None:
    traff = _finn_timer(timer, rad["klasse"], rad["dag"], rad["fag"])
    for treff in traff:
        treff["tema"] = " · ".join(x for x in (treff["tema"], rad["tema"]) if x)
        treff["lekse"] = " · ".join(x for x in (treff["lekse"], rad["lekse"]) if x)
        treff["type"] = rad["type"] or treff["type"]
    # En rad uten fag som bare er en frist, hører hjemme i fristlista – ikke i timeplanen.
    eget_kort = rad["tema"] or (rad["type"] and nokkel(rad["type"]) != "frist")
    if not traff and rad["dag"] and eget_kort:
        timer.append({
            "klasse": rad["klasse"], "dag": rad["dag"], "start": "", "slutt": "",
            "fag": rad["fag"], "laerer": "",
            "tema": rad["tema"] or rad["lekse"], "lekse": "", "type": rad["type"],
        })
        if rad["fag"]:
            merknader.append(
                f"Uke rad {nr}: fant ingen {rad['fag']}-time {rad['dag'].lower()} for {rad['klasse']}. "
                "Innholdet vises som eget kort den dagen."
            )
    if rad["lekse"] or rad["type"] in ("Prøve", "Innlevering", "Frist", "Tur"):
        oppgaver.append({
            "klasse": rad["klasse"],
            "fag": rad["fag"],
            "tekst": rad["lekse"] or rad["tema"] or rad["type"],
            "dag": rad["dag"] or rad["frist"],
            "frist": rad["frist"],
            "type": rad["type"],
        })


def _finn_timer(timer, klasse, dag, fag):
    """Finner timene uketeksten hører til: samme dag og fag, i riktig klasse.

    Står det «Alle» i klassefeltet, festes innholdet til timen i hver klasse."""
    if not dag or not fag:
        return []
    kandidater = [
        t for t in timer
        if t["dag"] == dag and nokkel(t["fag"]) == nokkel(fag)
        and (nokkel(t["klasse"]) == nokkel(klasse) or nokkel(klasse) == "alle" or nokkel(t["klasse"]) == "alle")
    ]
    if not kandidater:
        return []
    if nokkel(klasse) == "alle":
        per_klasse: dict[str, list] = {}
        for t in kandidater:
            per_klasse.setdefault(nokkel(t["klasse"]), []).append(t)
        return [_ledig(gruppe) for gruppe in per_klasse.values()]
    return [_ledig(kandidater)]


def _ledig(kandidater):
    """Første time uten innhold, ellers den første – to mattetimer samme dag
    får da hver sin tekst når det står to rader i Uke."""
    ledige = [t for t in kandidater if not t["tema"] and not t["lekse"]]
    return (ledige or kandidater)[0]
