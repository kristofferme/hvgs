"""Lager arbeidsboka: timeplanen som rutenett per klasse, og én fane per uke."""

from __future__ import annotations

import datetime as dt

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule, Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from .felles import (PROFILFARGE_STANDARD, TYPER, dagnavn, datospenn, farge_for, tekster,
                     tolk_sprak, uke_til_mandag)

SKRIFT = "Aptos Narrow"
BLEKK = "16212A"
GRA = "5B6A72"
ARK = "EFF2F0"
BAND = "F4F7F5"
KANT = Side(style="thin", color="D3D9D6")

# Fellesfaga de fleste videregående klasser har. Resten legger skolen selv inn.
STANDARDFAG = [
    "Norsk", "Matematikk 1T", "Matematikk 1P-Y", "Engelsk", "Naturfag",
    "Samfunnskunnskap", "Geografi", "Historie", "Religion og etikk", "Kroppsøving",
    "Spansk", "Yrkesfagleg fordjuping",
]
STANDARDKLASSER = ["1STA", "1STB", "2STA", "3STA", "1HSA", "1TIA"]
STANDARDOKTER = ["08:15–09:45", "10:00–11:30", "12:00–13:30", "13:40–15:10"]

OKTRADER = 8            # klokkeslettrader per klasse i Timeplan
INNHOLDSRADER = 400     # rader til tema og lekser
BESKJEDRADER = 120
LARERRADER = 120
UKER_I_LISTA = 46       # nedtrekket dekker et skoleår

def _kolonner(T):
    return {
        "uke": [(T["uke"], 26), (T["klasse"], 11), (T["kol_dag"], 12), (T["fag"], 22),
                (T["kol_tema"], 42), (T["kol_lekse"], 42), (T["kol_frist"], 12), (T["kol_type"], 14)],
        "beskjeder": [(T["uke"], 26), (T["klasse"], 11), (T["kol_tittel"], 26), (T["kol_beskjed"], 70)],
        "larere": [(T["klasse"], 11), (T["fag"], 26), (T["kol_larer"], 16)],
    }



def _tint(hex_farge: str, andel: float) -> str:
    h = hex_farge.lstrip("#")
    r, g, b = (int(h[i:i + 2], 16) for i in (0, 2, 4))
    bland = lambda v: int(round(v + (255 - v) * andel))
    return f"{bland(r):02X}{bland(g):02X}{bland(b):02X}"


def _skrift(**kwargs) -> Font:
    kwargs.setdefault("name", SKRIFT)
    kwargs.setdefault("size", 11)
    kwargs.setdefault("color", BLEKK)
    return Font(**kwargs)


def _tittel(ws, celle: str, tekst: str, storrelse: int = 11) -> None:
    ws[celle] = tekst
    ws[celle].font = _skrift(bold=True, size=storrelse)


def _tabell(ws, kolonner, hoderad: int, rader: int) -> None:
    """Hoderad med hvit skrift på svart, og ferdig formaterte rader under."""
    for i, (navn, bredde) in enumerate(kolonner, start=1):
        c = ws.cell(row=hoderad, column=i, value=navn)
        c.font = _skrift(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLEKK)
        c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws.column_dimensions[get_column_letter(i)].width = bredde
    ws.row_dimensions[hoderad].height = 26
    for rad in range(hoderad + 1, hoderad + 1 + rader):
        ws.row_dimensions[rad].height = 20
        for i, (_, bredde) in enumerate(kolonner, start=1):
            c = ws.cell(row=rad, column=i)
            c.border = Border(bottom=KANT, right=KANT)
            c.font = _skrift()
            c.alignment = Alignment(vertical="center", indent=1, wrap_text=bredde >= 40)


def _liste(ws, omrade: str, kilde: str, ledetekst: str) -> None:
    dv = DataValidation(type="list", formula1=kilde, allow_blank=True, showErrorMessage=False)
    dv.prompt = ledetekst
    dv.promptTitle = "Velg fra lista"
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(omrade)


def _fargekoder(ws, omrade: str, kolonne: str, forste_rad: int, fag: list[str]) -> None:
    """Farger fagcella, så uka kan leses på fargene alene."""
    brukte: dict[str, str] = {}
    for navn in fag:
        hex_ = farge_for(navn, brukte)
        ws.conditional_formatting.add(
            omrade,
            FormulaRule(
                formula=[f'EXACT(${kolonne}{forste_rad},"{navn}")'],
                fill=PatternFill("solid", bgColor=_tint(hex_, 0.86)),
                font=Font(name=SKRIFT, size=11, bold=True, color=hex_.lstrip("#")),
                stopIfTrue=True,
            ),
        )


def _klasseskille(ws, omrade: str, kolonne: str, forste_rad: int) -> None:
    """Skiller klassene: annenhver klasse tonet, og strek der klassen bytter."""
    ws.conditional_formatting.add(
        omrade,
        Rule(type="expression",
             formula=[f'AND(${kolonne}{forste_rad}<>"",ISODD(MATCH(${kolonne}{forste_rad},Klasser,0)))'],
             dxf=DifferentialStyle(fill=PatternFill("solid", bgColor=BAND))),
    )
    ws.conditional_formatting.add(
        omrade,
        Rule(type="expression",
             formula=[f'AND(${kolonne}{forste_rad}<>"",${kolonne}{forste_rad}<>${kolonne}{forste_rad - 1})'],
             dxf=DifferentialStyle(border=Border(top=Side(style="medium", color=BLEKK)))),
    )


def _ukeskille(ws, omrade: str, forste_rad: int) -> None:
    """Tykk strek der ukenummeret bytter, så uker skiller seg fra klasser."""
    ws.conditional_formatting.add(
        omrade,
        Rule(type="expression",
             formula=[f'AND($A{forste_rad}<>"",$A{forste_rad}<>$A{forste_rad - 1})'],
             dxf=DifferentialStyle(border=Border(top=Side(style="thick", color=BLEKK)))),
    )


def ukemerkelapp(mandag: dt.date) -> str:
    """«36 · 31. aug – 4. sep 2026» – slik uka står i nedtrekket."""
    return f"{mandag.isocalendar()[1]} · {datospenn(mandag, mandag + dt.timedelta(days=4))}"


def ukeliste(forste_dag: dt.date, antall: int = UKER_I_LISTA) -> list[str]:
    """Merkelapper for hele skoleåret, til nedtrekket i Uke-kolonnen."""
    start = forste_dag - dt.timedelta(weeks=6)
    return [ukemerkelapp(start + dt.timedelta(weeks=i)) for i in range(antall)]


# ── Ukeark ───────────────────────────────────────────────────────
def _timeplanrutenett(ws, klasser: list[str], fag: list[str], okter: list[str],
                      dager: list[str], timeplan: dict | None = None) -> int:
    """Ett rutenett per klasse: klokkeslett nedover, dager bortover.

    timeplan: {klasse: {dag: [fag per økt]}} fyller rutene. Gir siste rad."""
    timeplan = timeplan or {}
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 16
    for i in range(2, 7):
        ws.column_dimensions[get_column_letter(i)].width = 21

    klassefelt = DataValidation(type="list", formula1="Klasser", allow_blank=True, showErrorMessage=False)
    klassefelt.prompt, klassefelt.promptTitle, klassefelt.showInputMessage = (
        "Klassen dette rutenettet gjelder.", "Velg fra lista", True)
    fagfelt = DataValidation(type="list", formula1="Fagliste", allow_blank=True, showErrorMessage=False)
    fagfelt.prompt, fagfelt.promptTitle, fagfelt.showInputMessage = (
        "Faget i denne timen. Tomt betyr fri.", "Velg fra lista", True)
    ws.add_data_validation(klassefelt)
    ws.add_data_validation(fagfelt)

    rad = 1
    for klasse in klasser:
        c = ws.cell(row=rad, column=1, value=klasse)
        c.font = _skrift(bold=True, size=14, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLEKK)
        c.alignment = Alignment(vertical="center", horizontal="center")
        klassefelt.add(c)
        for i, dag in enumerate(dager, start=2):
            d = ws.cell(row=rad, column=i, value=dag)
            d.font = _skrift(bold=True)
            d.fill = PatternFill("solid", fgColor=ARK)
            d.alignment = Alignment(vertical="center", horizontal="center")
            d.border = Border(bottom=Side(style="medium", color=BLEKK))
        ws.row_dimensions[rad].height = 28

        for n in range(OKTRADER):
            r = rad + 1 + n
            ws.row_dimensions[r].height = 22
            t = ws.cell(row=r, column=1, value=okter[n] if n < len(okter) else None)
            t.font = Font(name=SKRIFT, size=10, color=GRA)
            t.alignment = Alignment(vertical="center", horizontal="center")
            t.border = Border(right=Side(style="medium", color=BLEKK), bottom=KANT)
            for i, dag in enumerate(dager, start=2):
                celle = ws.cell(row=r, column=i)
                celle.font = _skrift()
                celle.alignment = Alignment(vertical="center", indent=1)
                celle.border = Border(bottom=KANT, right=KANT)
                fylt = timeplan.get(klasse, {}).get(dag, [])
                if n < len(fylt) and fylt[n]:
                    celle.value = fylt[n]
        fagfelt.add(f"B{rad + 1}:F{rad + OKTRADER}")
        rad += OKTRADER + 2
    return rad


def _fargekoder_rutenett(ws, kolonner: str, rader: int, fag: list[str]) -> None:
    """Fargelegger hver dagkolonne i timeplanrutenettet for seg."""
    brukte: dict[str, str] = {}
    for navn in fag:
        hex_ = farge_for(navn, brukte)
        stil = DifferentialStyle(
            fill=PatternFill("solid", bgColor=_tint(hex_, 0.86)),
            font=Font(name=SKRIFT, size=11, bold=True, color=hex_.lstrip("#")),
        )
        for kol in kolonner:
            ws.conditional_formatting.add(
                f"{kol}1:{kol}{rader}",
                Rule(type="expression", formula=[f'EXACT(${kol}1,"{navn}")'], dxf=stil, stopIfTrue=True),
            )


def lag_arbeidsbok(sti, skole=None, uke=None, forste_dag=None, overskrift=None,
                   klasser=None, fag=None, okter=None, timeplan=None, innhold=None,
                   sprak="nb", profilfarge=PROFILFARGE_STANDARD, logo="") -> None:
    """Skriver en ferdig arbeidsbok. Uten innhold blir arkene tomme og klare."""
    sprak = tolk_sprak(sprak)
    T = tekster(sprak)
    KOL = _kolonner(T)
    skole = skole or ("Skulen" if sprak == "nn" else "Skolen")
    overskrift = overskrift or T["ukeplan"]
    klasser = klasser or STANDARDKLASSER
    fag = fag or STANDARDFAG
    okter = okter or STANDARDOKTER
    innhold = innhold or {}
    if forste_dag is None:
        i_dag = dt.date.today()
        forste_dag = i_dag - dt.timedelta(days=i_dag.weekday())
    if uke is None:
        uke = forste_dag.isocalendar()[1]
    uker = ukeliste(forste_dag)

    wb = Workbook()

    # ── Start her ────────────────────────────────────────────────
    start = wb.active
    start.title = T["ark_start"]
    start.sheet_view.showGridLines = False
    start.column_dimensions["A"].width = 4
    start.column_dimensions["B"].width = 104
    _tittel(start, "B2", T["ukeplan"], 22)
    linjer = [
        ("", "", ""),
        ("steg", "1  Oppsett – skole, ukenummer og datoen for mandagen. Legg inn klassene og fagene dine. Profilfarge og logo hører også hjemme her.",
                 "1  Oppsett – skule, vekenummer og datoen for måndagen. Legg inn klassane og faga dine. Profilfarge og logo høyrer også heime her."),
        ("steg", "2  Timeplan – ett rutenett per klasse. Fyll det ut én gang; det gjelder alle uker.",
                 "2  Timeplan – eitt rutenett per klasse. Fyll det ut éin gong; det gjeld alle veker."),
        ("steg", "3  Lærere – hvilken lærer faget har i hver klasse. Valgfritt.",
                 "3  Lærarar – kva lærar faget har i kvar klasse. Valfritt."),
        ("steg", "4  Uke – tema, lekser og frister. Velg uke i nedtrekket i første kolonne.",
                 "4  Veke – tema, lekser og fristar. Vel veke i nedtrekket i første kolonne."),
        ("steg", "5  Beskjeder – korte meldinger hjem, med samme ukevalg.",
                 "5  Meldingar – korte meldingar heim, med same vekeval."),
        ("", "", ""),
        ("steg", "Så kjører du:  python3 ukeplan.py bygg", "Så køyrer du:  python3 ukeplan.py bygg"),
        ("brod", "Du får én nettside med alle ukene. Elevene blar mellom dem med pilene eller piltastene.",
                 "Du får éi nettside med alle vekene. Elevane blar mellom dei med pilene eller piltastane."),
        ("", "", ""),
        ("mellom", "Når det blir mange uker", "Når det blir mange veker"),
        ("brod", "Uke-nedtrekket viser både ukenummer og datoer, så du slipper å telle. Én arbeidsbok tar hele skoleåret.",
                 "Veke-nedtrekket viser både vekenummer og datoar, så du slepp å telje. Éi arbeidsbok tek heile skuleåret."),
        ("brod", "Klikk på filterpila i Uke-kolonnen og hak av for én uke når du vil jobbe med bare den. Det er en tykk strek mellom ukene.",
                 "Klikk på filterpila i Veke-kolonnen og hak av for éi veke når du vil jobbe med berre den. Det er ein tjukk strek mellom vekene."),
        ("brod", "Skal du gjenta noe fra forrige uke: merk radene, kopier, lim inn nederst og bytt uke i nedtrekket.",
                 "Skal du gjenta noko frå førre veke: merk radene, kopier, lim inn nedst og byt veke i nedtrekket."),
        ("", "", ""),
        ("mellom", "Godt å vite", "Godt å vite"),
        ("brod", "Timeplanen gjentas i alle uker. Du fyller den ut én gang, ikke én gang per uke.",
                 "Timeplanen blir gjenteken i alle veker. Du fyller han ut éin gong, ikkje éin gong per veke."),
        ("brod", "Klasse, Dag, Fag, Frist og Type har nedtrekkslister. Klikk i cella og velg.",
                 "Klasse, Dag, Fag, Frist og Type har nedtrekkslister. Klikk i cella og vel."),
        ("brod", "Skriv «Alle» i Klasse-feltet når noe gjelder alle klassene.",
                 "Skriv «Alle» i Klasse-feltet når noko gjeld alle klassane."),
        ("brod", "Hver klasse får sin egen tone på ukearket, og det kommer en strek der klassen bytter.",
                 "Kvar klasse får sin eigen tone på vekearket, og det kjem ein strek der klassen byter."),
        ("brod", "Lekser og frister havner i «Å gjøre denne uka» på nettsiden, sortert etter frist.",
                 "Lekser og fristar hamnar i «Å gjere denne veka» på nettsida, sorterte etter frist."),
        ("brod", "Rader du ikke bruker, lar du stå tomme. Rekkefølgen spiller ingen rolle.",
                 "Rader du ikkje bruker, lèt du stå tomme. Rekkjefølgja spelar inga rolle."),
    ]
    rad = 3
    for slag, pa_bokmal, pa_nynorsk in linjer:
        tekst = pa_nynorsk if sprak == "nn" else pa_bokmal
        if tekst:
            c = start.cell(row=rad, column=2, value=tekst)
            if slag == "steg":
                c.font = _skrift(bold=True, size=12)
                start.row_dimensions[rad].height = 26
            elif slag == "mellom":
                c.font = _skrift(bold=True, color=GRA)
                start.row_dimensions[rad].height = 30
            else:
                c.font = _skrift(color=GRA)
                start.row_dimensions[rad].height = 30
            c.alignment = Alignment(vertical="center", wrap_text=True)
        rad += 1

    # ── Oppsett ──────────────────────────────────────────────────
    opp = wb.create_sheet(T["ark_oppsett"])
    opp.sheet_view.showGridLines = False
    for kol, bredde in (("A", 3), ("B", 24), ("C", 30), ("D", 34), ("E", 26), ("F", 14)):
        opp.column_dimensions[kol].width = bredde
    _tittel(opp, "B2", T["ark_oppsett"], 16)
    felt = [
        (T["skole"], skole, ""),
        (T["ukenummer"], uke, "← veka sida opnar på" if sprak == "nn" else "← uka nettsiden åpner på"),
        (T["mandag_i_uka"], forste_dag,
         "← herfrå reknar vi datoane i alle dei andre vekene" if sprak == "nn"
         else "← herfra regnes datoene i alle de andre ukene"),
        (T["overskrift"], overskrift, ""),
        (T["sprakfelt"], T["sprak"], "bokmål / nynorsk"),
        (T["profilfarge"], profilfarge, "← fargen skulen sin profil bruker" if sprak == "nn"
         else "← fargen skolens profil bruker"),
        (T["logo"], logo, "← ligg i same mappe som denne fila" if sprak == "nn"
         else "← ligger i samme mappe som denne fila"),
    ]
    for i, (navn, verdi, merknad) in enumerate(felt):
        r = 4 + i
        opp.cell(row=r, column=2, value=navn).font = _skrift(color=GRA)
        c = opp.cell(row=r, column=3, value=verdi)
        c.font = _skrift(bold=True, size=12)
        c.fill = PatternFill("solid", fgColor=ARK)
        c.border = Border(bottom=Side(style="medium", color=BLEKK))
        c.alignment = Alignment(vertical="center", indent=1)
        opp.row_dimensions[r].height = 22
        if merknad:
            opp.cell(row=r, column=4, value=merknad).font = _skrift(size=10, italic=True, color=GRA)
    opp["C6"].number_format = "dd.mm.yyyy"
    opp["C9"].fill = PatternFill("solid", fgColor=profilfarge.lstrip("#"))
    opp["C9"].font = _skrift(bold=True, size=12, color="FFFFFF")
    _liste(opp, "C8:C8", "Malformer", "bokmål eller nynorsk")

    _tittel(opp, "B12", T["klasser"])
    opp["B13"] = T["alle"]
    opp["B13"].font = _skrift(italic=True, color=GRA)
    opp["C13"] = ("← vel denne når noko gjeld alle" if sprak == "nn"
                  else "← velg denne når noe gjelder alle")
    opp["C13"].font = _skrift(size=10, italic=True, color=GRA)
    for i, k in enumerate(klasser):
        opp.cell(row=14 + i, column=2, value=k).font = _skrift()
    for r in range(13, 114):
        opp.cell(row=r, column=2).border = Border(bottom=KANT)

    _tittel(opp, "E12", T["fag"])
    _tittel(opp, "F12", T["farge"])
    brukte: dict[str, str] = {}
    for i, f in enumerate(fag):
        r = 13 + i
        opp.cell(row=r, column=5, value=f).font = _skrift()
        hex_ = farge_for(f, brukte)
        c = opp.cell(row=r, column=6, value=hex_)
        c.font = Font(name=SKRIFT, size=10, color=hex_.lstrip("#"))
        c.fill = PatternFill("solid", fgColor=_tint(hex_, 0.86))
        c.alignment = Alignment(horizontal="center")
    for r in range(13, 114):
        opp.cell(row=r, column=5).border = Border(bottom=KANT)

    # ── Navngitte områder ────────────────────────────────────────
    lister = wb.create_sheet(T["ark_lister"])
    lister["A1"] = "Dager"
    lister["B1"] = "Typer"
    dager = dagnavn(sprak)
    for i, d in enumerate(dager):
        lister.cell(row=2 + i, column=1, value=d)
    ekte_typer = TYPER[sprak]
    for i, t in enumerate(ekte_typer):
        lister.cell(row=2 + i, column=2, value=t)
    lister["F1"] = "Målformer"
    lister["F2"] = "bokmål"
    lister["F3"] = "nynorsk"
    lister["D1"] = T["uke"]
    for i, merkelapp in enumerate(uker):
        lister.cell(row=2 + i, column=4, value=merkelapp)
    lister.sheet_state = "hidden"

    A = T["ark_oppsett"]
    navn = {
        "Klasser": (f"OFFSET({A}!$B$13,0,0,MAX(2,COUNTA({A}!$B$13:$B$113)),1)"),
        "Fagliste": (f"OFFSET({A}!$E$13,0,0,MAX(1,COUNTA({A}!$E$13:$E$113)),1)"),
        "Malformer": f"{T['ark_lister']}!$F$2:$F$3",
        "Dager": f"{T['ark_lister']}!$A$2:$A${1 + len(dager)}",
        "Typer": f"{T['ark_lister']}!$B$2:$B${1 + len(ekte_typer)}",
        "Uker": f"{T['ark_lister']}!$D$2:$D${1 + len(uker)}",
    }
    for n, formel in navn.items():
        wb.defined_names.add(DefinedName(n, attr_text=formel))

    # ── Timeplan ─────────────────────────────────────────────────
    tp = wb.create_sheet(T["ark_timeplan"])
    siste = _timeplanrutenett(tp, klasser, fag, okter, dager, timeplan)
    _fargekoder_rutenett(tp, "BCDEF", siste, fag)

    # ── Lærere ───────────────────────────────────────────────────
    lr = wb.create_sheet(T["ark_larere"])
    lr.sheet_view.showGridLines = False
    _tabell(lr, KOL["larere"], 1, LARERRADER)
    lr.freeze_panes = "A2"
    lr.auto_filter.ref = f"A1:C{LARERRADER}"
    _liste(lr, f"A2:A{LARERRADER}", "Klasser", "Klassen. «Alle» gjelder alle klasser.")
    _liste(lr, f"B2:B{LARERRADER}", "Fagliste", "Faget.")
    _fargekoder(lr, f"B2:B{LARERRADER}", "B", 2, fag)
    _klasseskille(lr, f"A2:C{LARERRADER}", "A", 2)
    for r, rad in enumerate(innhold.get("Lærere", []), start=2):
        for k, verdi in enumerate(rad, start=1):
            if verdi not in (None, ""):
                lr.cell(row=r, column=k, value=verdi)

    # ── Uke ──────────────────────────────────────────────────────
    uk = wb.create_sheet(T["ark_uke"])
    uk.sheet_view.showGridLines = False
    _tabell(uk, KOL["uke"], 1, INNHOLDSRADER)
    uk.freeze_panes = "B2"
    uk.auto_filter.ref = f"A1:H{1 + INNHOLDSRADER}"
    sist = 1 + INNHOLDSRADER
    _liste(uk, f"A2:A{sist}", "Uker", "Velg uke. Lista viser ukenummer og datoer.")
    _liste(uk, f"B2:B{sist}", "Klasser", "Klassen dette gjelder. Velg «Alle» for alle klasser.")
    _liste(uk, f"C2:C{sist}", "Dager", "Dagen innholdet hører til.")
    _liste(uk, f"D2:D{sist}", "Fagliste", "Faget. Må stemme med timeplanen for å havne i riktig time.")
    _liste(uk, f"G2:G{sist}", "Dager", "Når må det være gjort? La stå tomt om det ikke har frist.")
    _liste(uk, f"H2:H{sist}", "Typer", "Merk prøver, innleveringer og turer.")
    _fargekoder(uk, f"D2:D{sist}", "D", 2, fag)
    _klasseskille(uk, f"A2:H{sist}", "B", 2)
    _ukeskille(uk, f"A2:H{sist}", 2)

    # ── Beskjeder ────────────────────────────────────────────────
    be = wb.create_sheet(T["ark_beskjeder"])
    be.sheet_view.showGridLines = False
    _tabell(be, KOL["beskjeder"], 1, BESKJEDRADER)
    be.freeze_panes = "B2"
    be.auto_filter.ref = f"A1:D{1 + BESKJEDRADER}"
    _liste(be, f"A2:A{1 + BESKJEDRADER}", "Uker", "Velg uke.")
    _liste(be, f"B2:B{1 + BESKJEDRADER}", "Klasser", "Klassen beskjeden gjelder. «Alle» går til alle.")
    _klasseskille(be, f"A2:D{1 + BESKJEDRADER}", "B", 2)
    _ukeskille(be, f"A2:D{1 + BESKJEDRADER}", 2)

    for arknavn in (T["ark_uke"], T["ark_beskjeder"]):
        for r, rad in enumerate(innhold.get(arknavn, innhold.get({T["ark_uke"]: "Uke", T["ark_beskjeder"]: "Beskjeder"}[arknavn], [])), start=2):
            for k, verdi in enumerate(rad, start=1):
                if k == 1 and isinstance(verdi, int):
                    verdi = ukemerkelapp(uke_til_mandag(verdi, forste_dag))
                if verdi not in (None, ""):
                    wb[arknavn].cell(row=r, column=k, value=verdi)

    wb.active = wb.sheetnames.index(T["ark_start"])
    wb.save(sti)
