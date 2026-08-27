"""Lager arbeidsboka: elevlister per klasse, innmeldingsark og tiltaksark."""

from __future__ import annotations

import re

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule, Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from .felles import (LYS, LYSFARGER, OMRADER, PROFILFARGE_STANDARD, STANDARDKLASSER,
                     lysnavn, statusnavn, tekster, tint, tolk_sprak)

SKRIFT = "Aptos Narrow"
BLEKK = "16212A"
GRA = "5B6A72"
ARK = "EFF2F0"
BAND = "F4F7F5"
KANT = Side(style="thin", color="D3D9D6")

ELEVRADER = 40          # elever per klasse
INNMELDINGSRADER = 700
TILTAKSRADER = 250
KLASSERADER = 60
OMRADERADER = 24
MOTERADER = 20


def _tint(hex_farge: str, andel: float) -> str:
    return tint(hex_farge, andel)


def _skrift(**kwargs) -> Font:
    kwargs.setdefault("name", SKRIFT)
    kwargs.setdefault("size", 11)
    kwargs.setdefault("color", BLEKK)
    return Font(**kwargs)


def _tittel(ws, celle: str, tekst: str, storrelse: int = 11) -> None:
    ws[celle] = tekst
    ws[celle].font = _skrift(bold=True, size=storrelse)


def _tabell(ws, kolonner, hoderad: int, rader: int) -> None:
    """Hoderad med hvit skrift på svart, og ferdig formaterte rader under."""
    for i, (navn, bredde) in enumerate(kolonner, start=1):
        c = ws.cell(row=hoderad, column=i, value=navn)
        c.font = _skrift(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLEKK)
        c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws.column_dimensions[get_column_letter(i)].width = bredde
    ws.row_dimensions[hoderad].height = 26
    for rad in range(hoderad + 1, hoderad + 1 + rader):
        ws.row_dimensions[rad].height = 20
        for i, (_, bredde) in enumerate(kolonner, start=1):
            c = ws.cell(row=rad, column=i)
            c.border = Border(bottom=KANT, right=KANT)
            c.font = _skrift()
            c.alignment = Alignment(vertical="center", indent=1, wrap_text=bredde >= 50)


def _liste(ws, omrade: str, kilde: str, ledetekst: str) -> None:
    dv = DataValidation(type="list", formula1=kilde, allow_blank=True, showErrorMessage=False)
    dv.prompt = ledetekst
    dv.promptTitle = "Velg fra lista"
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(omrade)


def _lysfarger(ws, sterkt: str, svakt: str, kolonne: str, forste_rad: int) -> None:
    """Farger lyscella kraftig og resten av raden svakt, så arket kan leses på farge alene."""
    for i, (_, kode) in enumerate(LYS["nb"]):
        navn = {LYS[s][i][0] for s in ("nb", "nn")}
        prove = "OR(" + ",".join(f'EXACT(${kolonne}{forste_rad},"{n}")' for n in sorted(navn)) + ")"
        hex_ = LYSFARGER[kode]
        ws.conditional_formatting.add(
            sterkt,
            FormulaRule(formula=[prove],
                        fill=PatternFill("solid", bgColor=_tint(hex_, 0.72)),
                        font=Font(name=SKRIFT, size=11, bold=True, color=hex_.lstrip("#")),
                        stopIfTrue=True))
        ws.conditional_formatting.add(
            svakt,
            FormulaRule(formula=[prove],
                        fill=PatternFill("solid", bgColor=_tint(hex_, 0.93)),
                        stopIfTrue=True))


def _skille(ws, omrade: str, kolonne: str, forste_rad: int, tykkelse: str = "medium") -> None:
    """Strek der verdien i kolonnen bytter – skiller møter fra hverandre, og klasser."""
    ws.conditional_formatting.add(
        omrade,
        Rule(type="expression",
             formula=[f'AND(${kolonne}{forste_rad}<>"",'
                      f'${kolonne}{forste_rad}<>${kolonne}{forste_rad - 1})'],
             dxf=DifferentialStyle(border=Border(top=Side(style=tykkelse, color=BLEKK)))))


def omradenavn(klasse: str) -> str:
    """«2AKV/FF» → «ELEV_2AKV_FF». Navngitte områder tåler ikke / og mellomrom."""
    return "ELEV_" + re.sub(r"[^A-Za-z0-9_]", "_", klasse)


def _elevark(ws, klasser: list[str], elever: dict, T) -> dict:
    """Én kolonne per klasse. Elev-nedtrekket følger klassen som er valgt."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    omrader = {}
    for i, klasse in enumerate(klasser, start=1):
        bokstav = get_column_letter(i)
        ws.column_dimensions[bokstav].width = 26
        c = ws.cell(row=1, column=i, value=klasse)
        c.font = _skrift(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLEKK)
        c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        for r in range(2, 2 + ELEVRADER):
            celle = ws.cell(row=r, column=i)
            celle.font = _skrift()
            celle.alignment = Alignment(vertical="center", indent=1)
            celle.border = Border(bottom=KANT, right=KANT)
        for r, navn in enumerate(elever.get(klasse, []), start=2):
            ws.cell(row=r, column=i, value=navn)
        omrader[omradenavn(klasse)] = (
            f"OFFSET('{T['elevar_ark']}'!${bokstav}$2,0,0,"
            f"MAX(1,COUNTA('{T['elevar_ark']}'!${bokstav}$2:${bokstav}${1 + ELEVRADER})),1)")
    ws.row_dimensions[1].height = 26
    return omrader


def _elevformel(kolonne: str, rad: int) -> str:
    """Nedtrekket for Elev peker på klassens egen liste."""
    b = f"${kolonne}{rad}"
    return (f'IFERROR(INDIRECT("ELEV_"&SUBSTITUTE(SUBSTITUTE(SUBSTITUTE({b},"/","_")," ","_"),'
            f'"-","_")),VelKlasse)')


def lag_arbeidsbok(sti, skole=None, skolear="", klasser=None, elever=None, omrader=None,
                   moter=None, innmeldingar=None, tiltak=None, sprak="nb",
                   profilfarge=PROFILFARGE_STANDARD, logo="") -> None:
    """Skriver en ferdig arbeidsbok. Uten innhold blir arkene tomme og klare."""
    sprak = tolk_sprak(sprak)
    T = tekster(sprak)
    skole = skole or ("Skulen" if sprak == "nn" else "Skolen")
    klasser = klasser or STANDARDKLASSER
    elever = elever or {}
    omrader = omrader or OMRADER[sprak]
    moter = moter or []
    lysval = lysnavn(sprak)
    statusval = statusnavn(sprak)

    wb = Workbook()

    # ── Start her ────────────────────────────────────────────────
    start = wb.active
    start.title = T["ark_start"]
    start.sheet_view.showGridLines = False
    start.column_dimensions["A"].width = 4
    start.column_dimensions["B"].width = 106
    _tittel(start, "B2", T["trafikklys"], 22)
    linjer = [
        ("", "", ""),
        ("brod", "Hver lærer melder inn det som ikke er grønt. Møtet får ett rutenett per klasse: elevene nedover, områdene bortover, sterkeste lys i hver rute.",
                 "Kvar lærar melder inn det som ikkje er grønt. Møtet får eitt rutenett per klasse: elevane nedover, områda bortover, sterkaste lys i kvar rute."),
        ("", "", ""),
        ("steg", "1  Oppsett – skole, klasser, møtedatoer og hvilke områder dere setter lys på.",
                 "1  Oppsett – skule, klassar, møtedatoar og kva område de set lys på."),
        ("steg", "2  Elever – én kolonne per klasse med navna. Da vet Innmelding hvem som finnes.",
                 "2  Elevar – éi kolonne per klasse med namna. Då veit Innmelding kven som finst."),
        ("steg", "3  Innmelding – én rad per ting du vil melde inn. Alt har nedtrekksliste.",
                 "3  Innmelding – éi rad per ting du vil melde inn. Alt har nedtrekksliste."),
        ("steg", "4  Tiltak – det møtet blir enige om: hva, hvem og til når.",
                 "4  Tiltak – det møtet blir samde om: kva, kven og til når."),
        ("", "", ""),
        ("steg", "Så kjører du:  python3 trafikklys.py bygg", "Så køyrer du:  python3 trafikklys.py bygg"),
        ("brod", "Det gir Elevstatus.html – møtevisningen. Den åpner du i nettleseren og deler på skjermen i møtet.",
                 "Det gir Elevstatus.html – møtevisinga. Den opnar du i nettlesaren og deler på skjermen i møtet."),
        ("", "", ""),
        ("mellom", "Hva lysene betyr", "Kva lysa tyder"),
        ("brod", "Grønt – alt er i orden. Trenger ikke meldes inn, men du kan gjøre det om du vil bekrefte at du har sett etter.",
                 "Grønt – alt er i orden. Treng ikkje meldast inn, men du kan gjere det om du vil stadfeste at du har sett etter."),
        ("brod", "Gult – noe å følge med på. Kontaktlærer og faglærer håndterer det i klassen, men møtet skal vite om det.",
                 "Gult – noko å følgje med på. Kontaktlærar og faglærar handterer det i klassen, men møtet skal vite om det."),
        ("brod", "Rødt – må tas opp i elevstatusmøtet. Skal ende i et tiltak med en ansvarlig og en frist.",
                 "Raudt – må takast opp i elevstatusmøtet. Skal ende i eit tiltak med ein ansvarleg og ein frist."),
        ("", "", ""),
        ("mellom", "Slik skriver du merknaden", "Slik skriv du merknaden"),
        ("brod", "Skriv det du har observert, kort og konkret: «Ikke levert de tre siste innleveringene» framfor «umotivert».",
                 "Skriv det du har observert, kort og konkret: «Ikkje levert dei tre siste innleveringane» framfor «umotivert»."),
        ("brod", "Ikke skriv helseopplysninger, diagnoser eller opplysninger om familien. Det hører hjemme i elevmappa, ikke her.",
                 "Ikkje skriv helseopplysningar, diagnosar eller opplysningar om familien. Det høyrer heime i elevmappa, ikkje her."),
        ("brod", "Denne fila inneholder personopplysninger. Den skal ligge i Teams der bare de som skal ha tilgang har det – aldri på en åpen nettadresse.",
                 "Denne fila inneheld personopplysningar. Ho skal liggje i Teams der berre dei som skal ha tilgang har det – aldri på ei open nettadresse."),
        ("", "", ""),
        ("mellom", "Godt å vite", "Godt å vite"),
        ("brod", "Elev-nedtrekket viser bare elevene i klassen du har valgt. Velg klasse først.",
                 "Elev-nedtrekket viser berre elevane i klassen du har valt. Vel klasse først."),
        ("brod", "Møte-kolonnen gjør at én arbeidsbok tar hele året. Møtevisningen viser piler for det som er nytt eller endret siden forrige møte.",
                 "Møte-kolonnen gjer at éi arbeidsbok tek heile året. Møtevisinga viser piler for det som er nytt eller endra sidan førre møte."),
        ("brod", "Klikk på filterpila i Møte-kolonnen når du vil se ett møte om gangen. Rader du ikke bruker, lar du stå tomme.",
                 "Klikk på filterpila i Møte-kolonnen når du vil sjå eitt møte om gongen. Rader du ikkje bruker, lèt du stå tomme."),
    ]
    rad = 3
    for slag, pa_bokmal, pa_nynorsk in linjer:
        tekst = pa_nynorsk if sprak == "nn" else pa_bokmal
        if tekst:
            c = start.cell(row=rad, column=2, value=tekst)
            if slag == "steg":
                c.font = _skrift(bold=True, size=12)
                start.row_dimensions[rad].height = 26
            elif slag == "mellom":
                c.font = _skrift(bold=True, color=GRA)
                start.row_dimensions[rad].height = 30
            else:
                c.font = _skrift(color=GRA)
                start.row_dimensions[rad].height = 30
            c.alignment = Alignment(vertical="center", wrap_text=True)
        rad += 1

    # ── Oppsett ──────────────────────────────────────────────────
    opp = wb.create_sheet(T["ark_oppsett"])
    opp.sheet_view.showGridLines = False
    for kol, bredde in (("A", 3), ("B", 22), ("C", 30), ("D", 26), ("E", 3), ("F", 28), ("G", 54)):
        opp.column_dimensions[kol].width = bredde
    _tittel(opp, "B2", T["ark_oppsett"], 16)
    felt = [
        (T["skole"], skole, ""),
        (T["skoleaar"], skolear, ""),
        (T["sprakfelt"], T["sprak"], "bokmål / nynorsk"),
        (T["profilfarge"], profilfarge,
         "← fargen skulen sin profil bruker" if sprak == "nn" else "← fargen skolens profil bruker"),
        (T["logo"], logo,
         "← ligg i same mappe som denne fila" if sprak == "nn" else "← ligger i samme mappe som denne fila"),
    ]
    for i, (navn, verdi, merknad) in enumerate(felt):
        r = 4 + i
        opp.cell(row=r, column=2, value=navn).font = _skrift(color=GRA)
        c = opp.cell(row=r, column=3, value=verdi)
        c.font = _skrift(bold=True, size=12)
        c.fill = PatternFill("solid", fgColor=ARK)
        c.border = Border(bottom=Side(style="medium", color=BLEKK))
        c.alignment = Alignment(vertical="center", indent=1)
        opp.row_dimensions[r].height = 22
        if merknad:
            opp.cell(row=r, column=4, value=merknad).font = _skrift(size=10, italic=True, color=GRA)
    opp["C7"].fill = PatternFill("solid", fgColor=profilfarge.lstrip("#"))
    opp["C7"].font = _skrift(bold=True, size=12, color="FFFFFF")
    _liste(opp, "C6:C6", "Malformer", "bokmål eller nynorsk")

    for celle, tittel in (("B11", T["klasser"]), ("C11", T["moter"]), ("D11", T["dato"]),
                          ("F11", T["omrader"]), ("G11", T["forklaring"])):
        _tittel(opp, celle, tittel)
    for i, k in enumerate(klasser):
        opp.cell(row=12 + i, column=2, value=k).font = _skrift()
    for i, (navn, dato) in enumerate(moter):
        opp.cell(row=12 + i, column=3, value=navn).font = _skrift()
        c = opp.cell(row=12 + i, column=4, value=dato)
        c.font = _skrift()
        c.number_format = "dd.mm.yyyy"
    for i, (navn, forklaring) in enumerate(omrader):
        opp.cell(row=12 + i, column=6, value=navn).font = _skrift()
        opp.cell(row=12 + i, column=7, value=forklaring).font = _skrift(size=10, color=GRA)
    for r in range(12, 12 + KLASSERADER):
        opp.cell(row=r, column=2).border = Border(bottom=KANT)
    for r in range(12, 12 + MOTERADER):
        opp.cell(row=r, column=3).border = Border(bottom=KANT)
        opp.cell(row=r, column=4).border = Border(bottom=KANT)
        opp.cell(row=r, column=4).number_format = "dd.mm.yyyy"
    for r in range(12, 12 + OMRADERADER):
        opp.cell(row=r, column=6).border = Border(bottom=KANT)
        opp.cell(row=r, column=7).border = Border(bottom=KANT)
    opp.cell(row=12 + MOTERADER + 1, column=3,
             value=("← eitt møte per rad, nyaste nedst" if sprak == "nn"
                    else "← ett møte per rad, nyeste nederst")).font = _skrift(size=10, italic=True, color=GRA)

    # ── Lister (skjult) ──────────────────────────────────────────
    lister = wb.create_sheet(T["ark_lister"])
    lister["A1"] = T["lys"]
    for i, navn in enumerate(lysval):
        lister.cell(row=2 + i, column=1, value=navn)
    lister["B1"] = T["status"]
    for i, navn in enumerate(statusval):
        lister.cell(row=2 + i, column=2, value=navn)
    lister["D1"] = "Målformer"
    lister["D2"] = "bokmål"
    lister["D3"] = "nynorsk"
    lister["F1"] = T["elev"]
    lister["F2"] = "← vel klasse først" if sprak == "nn" else "← velg klasse først"
    lister.sheet_state = "hidden"

    A = T["ark_oppsett"]
    navngitte = {
        "Klasser": f"OFFSET({A}!$B$12,0,0,MAX(1,COUNTA({A}!$B$12:$B${11 + KLASSERADER})),1)",
        "Moter": f"OFFSET({A}!$C$12,0,0,MAX(1,COUNTA({A}!$C$12:$C${11 + MOTERADER})),1)",
        "Omrader": f"OFFSET({A}!$F$12,0,0,MAX(1,COUNTA({A}!$F$12:$F${11 + OMRADERADER})),1)",
        "Lysliste": f"'{T['ark_lister']}'!$A$2:$A${1 + len(lysval)}",
        "Statusar": f"'{T['ark_lister']}'!$B$2:$B${1 + len(statusval)}",
        "Malformer": f"'{T['ark_lister']}'!$D$2:$D$3",
        "VelKlasse": f"'{T['ark_lister']}'!$F$2",
    }

    # ── Elever ───────────────────────────────────────────────────
    elevark = wb.create_sheet(T["elevar_ark"])
    navngitte.update(_elevark(elevark, klasser, elever, T))
    for n, formel in navngitte.items():
        wb.defined_names.add(DefinedName(n, attr_text=formel))

    # ── Innmelding ───────────────────────────────────────────────
    inn = wb.create_sheet(T["ark_innmelding"])
    inn.sheet_view.showGridLines = False
    kolonner = [(T["mote"], 26), (T["klasse"], 11), (T["elev"], 26), (T["omrade"], 26),
                (T["lys"], 12), (T["merknad"], 64), (T["larer"], 20)]
    _tabell(inn, kolonner, 1, INNMELDINGSRADER)
    inn.freeze_panes = "B2"
    sist = 1 + INNMELDINGSRADER
    inn.auto_filter.ref = f"A1:G{sist}"
    _liste(inn, f"A2:A{sist}", "Moter", "Hvilket møte gjelder innmeldingen? Lista står i Oppsett.")
    _liste(inn, f"B2:B{sist}", "Klasser", "Klassen eleven går i.")
    _liste(inn, f"C2:C{sist}", _elevformel("B", 2), "Eleven. Lista viser elevene i klassen du valgte.")
    _liste(inn, f"D2:D{sist}", "Omrader", "Hva gjelder det? Områdene står i Oppsett.")
    _liste(inn, f"E2:E{sist}", "Lysliste", "Grønt, gult eller rødt.")
    _lysfarger(inn, f"E2:E{sist}", f"A2:D{sist} F2:G{sist}", "E", 2)
    _skille(inn, f"A2:G{sist}", "B", 2, "medium")
    _skille(inn, f"A2:G{sist}", "A", 2, "thick")

    # ── Tiltak ───────────────────────────────────────────────────
    til = wb.create_sheet(T["ark_tiltak"])
    til.sheet_view.showGridLines = False
    tkol = [(T["mote"], 26), (T["klasse"], 11), (T["elev"], 26), (T["omrade"], 24),
            (T["tiltak"], 56), (T["ansvarleg"], 22), (T["frist"], 14), (T["status"], 18)]
    _tabell(til, tkol, 1, TILTAKSRADER)
    til.freeze_panes = "B2"
    tsist = 1 + TILTAKSRADER
    til.auto_filter.ref = f"A1:H{tsist}"
    _liste(til, f"A2:A{tsist}", "Moter", "Møtet tiltaket ble bestemt på.")
    _liste(til, f"B2:B{tsist}", "Klasser", "Klassen eleven går i.")
    _liste(til, f"C2:C{tsist}", _elevformel("B", 2), "Eleven. Velg klasse først.")
    _liste(til, f"D2:D{tsist}", "Omrader", "Området tiltaket svarer på.")
    _liste(til, f"H2:H{tsist}", "Statusar", "Hvor langt er tiltaket kommet?")
    for r in range(2, tsist + 1):
        til.cell(row=r, column=7).number_format = "dd.mm.yyyy"
    avslutta = "OR(" + ",".join(f'EXACT($H2,"{n}")' for n in
                                sorted({"Avsluttet", "Avslutta"})) + ")"
    # Frist som er gått ut, på et tiltak som ikke er avsluttet.
    til.conditional_formatting.add(
        f"G2:G{tsist}",
        FormulaRule(formula=[f'AND($G2<>"",$G2<TODAY(),NOT({avslutta}))'],
                    fill=PatternFill("solid", bgColor=_tint(LYSFARGER["rod"], 0.72)),
                    font=Font(name=SKRIFT, size=11, bold=True, color=LYSFARGER["rod"].lstrip("#")),
                    stopIfTrue=True))
    til.conditional_formatting.add(
        f"A2:H{tsist}",
        FormulaRule(formula=[avslutta],
                    font=Font(name=SKRIFT, size=11, color=GRA, italic=True),
                    stopIfTrue=True))
    _skille(til, f"A2:H{tsist}", "B", 2, "medium")
    _skille(til, f"A2:H{tsist}", "A", 2, "thick")

    # ── Innhold ──────────────────────────────────────────────────
    for arknavn, rader in ((T["ark_innmelding"], innmeldingar or []),
                           (T["ark_tiltak"], tiltak or [])):
        for r, rad in enumerate(rader, start=2):
            for k, verdi in enumerate(rad, start=1):
                if verdi not in (None, ""):
                    wb[arknavn].cell(row=r, column=k, value=verdi)

    # Hjelpelista hører hjemme bakerst, bak arka folk faktisk bruker.
    wb.move_sheet(T["ark_lister"],
                  offset=len(wb.sheetnames) - 1 - wb.sheetnames.index(T["ark_lister"]))
    wb.active = wb.sheetnames.index(T["ark_start"])
    wb.save(sti)
